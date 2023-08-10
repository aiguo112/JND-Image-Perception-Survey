import tkinter
import customtkinter as tk
# import tkinter as tk
from tkinter import filedialog, StringVar
from PIL import Image, ImageTk
import os
import openpyxl
import pandas as pd


class JND_Survey(tk.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("1280x720")
        tk.set_appearance_mode("light")  # Modes: system (default), light, dark
        tk.set_default_color_theme("dark-blue")  # Themes: blue (default), dark-blue, green
        self.title("FlowTile JND Subjective Test")
        self.geometry("1280x720")
        self.configure(background="white")
        self.iconbitmap('Picture1.ico')
        radio_font = 15
        quality_var: StringVar

        # Initialize the image index
        self.image_index = 0
        self.User_ID = ""
        self.File_name="existing_file.xlsx"

        self.Quiz_count = 0
        # Initialize the list of image filenames
        self.image_filenames = []
        self.quality_var = tk.StringVar()
        self.quality_var.set(None)

        # create all of the main containers
        top_frame = tk.CTkFrame(self, bg_color="transparent", width=450, height=50)
        center = tk.CTkFrame(self, bg_color="transparent", width=50, height=40)
        btm_frame = tk.CTkFrame(self, bg_color="transparent", width=450, height=45)
       # btm_frame2 = tk.CTkFrame(self, bg_color="transparent", width=450, height=60)
        self.Image_Pair_No = []
        self.Image_Score = []
        # layout all of the main containers
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        top_frame.grid(row=0)
        center.grid(row=1, sticky="nsew")
        btm_frame.grid(row=3, sticky="ew")
        #btm_frame2.grid(row=4, sticky="ew")
        self.ctr_mid = tk.CTkFrame(center, bg_color="transparent", width=500, height=190)
        self.ctr_right = tk.CTkFrame(center, bg_color="transparent", width=200, height=190)

        # create the widgets for the top frame
        model_label = tk.CTkLabel(top_frame, text="FlowTile JND Subjective Test",
                                  font=("courier", 30, "bold", "underline"), text_color="#3A7EBF")
        self.User_ID_label = tk.CTkLabel(top_frame, text='UserID:', font=("courier", radio_font, "bold"),
                                         text_color="red")
        self.quiz_label = tk.CTkLabel(top_frame, text='(Image Pair No 6)', font=("courier", radio_font, "bold"),
                                      text_color="red")
        self.Radio_label = tk.CTkLabel(self.ctr_right, text='Please Select One Option: ',
                                       font=("courier", 20 , "underline"))
        self.instructions_label = tk.CTkLabel(btm_frame,
                                              text="Welcome to our image quality comparison survey! We appreciate you "
                                                   "taking the time to participate. \nIn this survey, you will be asked "
                                                   "to compare the quality of two images and select the one that you "
                                                   "believe is of higher quality. "
                                                   "\n"
                                                   "\n"
                                                   "\nTo complete the survey, please follow these instructions:"
                                                   "\n"
                                                   "\n✔Read the instructions on each screen carefully."
                                                   "\n✔When prompted, compare the two images displayed and select the one "
                                                   "that you believe is of higher quality. "
                                                   "\n✔Repeat this process for each pair of images in the survey."
                                                   "\n✔Once you have completed the survey, click the submit button to "
                                                   "submit your responses. "
                                                   "\n✔Thank you for your participation! Your feedback is valuable to us "
                                                   "and will help us improve our image quality standards.",
                                              font=("courier", 12), justify=tk.LEFT)

        self.checkbox_1 = tk.CTkCheckBox(btm_frame, text="I agree with the terms and conditions",
                                         font=("courier", 12))

        # layout the widgets in the top frame
        model_label.grid(row=0, column=0, columnspan=3, sticky="NESW")
        model_label.grid_rowconfigure(1, weight=2)
        model_label.grid_columnconfigure(1, weight=2)
        self.User_ID_label.grid(row=1, column=0)
        self.quiz_label.grid(row=1, column=2)
        self.Radio_label.pack(anchor=tk.W, pady=5, padx=10)
        self.instructions_label.pack(anchor=tk.W, pady=15, padx=2)
        self.checkbox_1.pack(anchor=tk.W, pady=15, padx=2)

        # create the center widgets
        center.grid_rowconfigure(0, weight=1)
        center.grid_columnconfigure(1, weight=1)

        self.radio_button1 = tk.CTkRadioButton(self.ctr_right, text="The left image has visible defects.",
                                               variable=self.quality_var, value="-3", font=("courier", radio_font))
        self.radio_button2 = tk.CTkRadioButton(self.ctr_right, text="The left image has inaccurate colors.",
                                               variable=self.quality_var, value="-2"
                                               , font=("courier", radio_font))
        self.radio_button3 = tk.CTkRadioButton(self.ctr_right, text="The left image has low contrast.",
                                               variable=self.quality_var, value="-1", font=("courier", radio_font))
        self.radio_button4 = tk.CTkRadioButton(self.ctr_right, text="Both images look the same.",
                                               variable=self.quality_var,
                                               value="0", font=("courier", radio_font))
        self.radio_button5 = tk.CTkRadioButton(self.ctr_right, text="The right image has no visible defects.",
                                               variable=self.quality_var, value="+1", font=("courier", radio_font))
        self.radio_button6 = tk.CTkRadioButton(self.ctr_right, text="The right image has better colors.",
                                               variable=self.quality_var, value="+2", font=("courier", radio_font))
        self.radio_button7 = tk.CTkRadioButton(self.ctr_right, text="The right image has better contrast.",
                                               variable=self.quality_var, value="+3", font=("courier", radio_font))
        self.radio_button1.pack(anchor=tk.W, pady=5, padx=10)
        self.radio_button2.pack(anchor=tk.W, pady=5, padx=10)
        self.radio_button3.pack(anchor=tk.W, pady=5, padx=10)
        self.radio_button4.pack(anchor=tk.W, pady=5, padx=10)
        self.radio_button5.pack(anchor=tk.W, pady=5, padx=10)
        self.radio_button6.pack(anchor=tk.W, pady=5, padx=10)
        self.radio_button7.pack(anchor=tk.W, pady=5, padx=10)

        self.button1_next = tk.CTkButton(self.ctr_right, text="Next", command=self.on_Next_button_click, bg_color="#2596be",
                                         font=("courier", 15, "bold"))


        self.button2_load = tk.CTkButton(self.ctr_right, text="Load Images", command=self.on_folder_button_click,
                                         bg_color="#2596be",
                                         font=("courier", 15, "bold"))
        self.button2_load.pack(pady=10, padx=5)

        self.button3_submit = tk.CTkButton(self.ctr_right, text="Submit Survey", command=self.save_to_excel,
                                           bg_color="#2596be",
                                           font=("courier", 15, "bold"))

        self.button4_quit = tk.CTkButton(self.ctr_right, text="Quit", command=self.destroy, bg_color="#2596be",
                                         font=("courier", 15, "bold"))
        self.button4_quit.pack(pady=10, padx=5)

        self.submit_label = tk.CTkLabel(self.ctr_right, text='✔ Submitted. Thank You', font=("courier", radio_font, "bold"),
                                        text_color="green")

        self.bottom_start = tk.CTkButton(btm_frame, text="Start", command=self.on_submit, bg_color="#2596be",
                                         font=("courier", 15, "bold"))
        self.bottom_start.pack(anchor=tk.W, pady=10, padx=2)

        # Create the labels for the images
        self.image_label1 = tk.CTkLabel(self.ctr_mid, text="  ")
        self.image_label1.pack(side=tk.LEFT)
        self.image_label2 = tk.CTkLabel(self.ctr_mid, text="  ")
        self.image_label2.pack(side=tk.LEFT)


        self.ctr_mid.pack_forget()
        self.ctr_right.pack_forget()


    def on_folder_button_click(self):


        # self.open_new_window()
        # Initialize the list of image filenames
        self.Image_Score.clear()
        # Open a file dialog to select the folder
        folder_path = filedialog.askdirectory()

        # Get a list of all the image filenames in the folder
        self.image_filenames = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".jpg")]

        # Sort the list of filenames alphabetically
        self.image_filenames.sort()
        # Load the first two images
        self.load_images()
        self.button1_next.pack(pady=10, padx=5)
        self.button2_load.pack_forget()
        self.button3_submit.pack(pady=10, padx=5)
        self.button4_quit.pack_forget()

    def load_images(self):

        # Get the filenames of the next two images
        image1_filename = self.image_filenames[self.image_index]
        image2_filename = self.image_filenames[self.image_index + 1]

        # Open the images and resize them
        image1 = Image.open(image1_filename)
        image1 = image1.resize((640, 320), Image.ANTIALIAS)
        image1 = ImageTk.PhotoImage(image1)

        image2 = Image.open(image2_filename)
        image2 = image2.resize((640, 320), Image.ANTIALIAS)
        image2 = ImageTk.PhotoImage(image2)

        # Update the labels with the new images
        self.image_label1.configure(image=image1, padx=10, pady=20)
        self.image_label1.image = image1
        self.image_label2.configure(image=image2, padx=10, pady=20)
        self.image_label2.image = image2

    def on_Next_button_click(self):
        # Increment the image index
        self.Quiz_count = self.Quiz_count + 1
        self.image_index += 2
        selected_answer = self.quality_var.get()
        # answers.append(selected_answer)
        self.quiz_label.configure(text=f"(Image Pair No.{self.Quiz_count})")
        print(self.Quiz_count)
        print(self.quality_var.get())

        # Clear the radio buttons for the next question
        if self.quality_var.get():
            self.add_value(self.Quiz_count, self.quality_var.get())
            self.quality_var.set(None)

        if self.image_index < len(self.image_filenames):
            image1_filename = self.image_filenames[self.image_index]
            self.load_images()
        else:
            print("Thank you, the folder has no more image pairs"
                  "")
            return

    def add_value(self, imagepair, imagescore):
        self.Image_Pair_No.append(imagepair)
        self.Image_Score.append(imagescore)

    def save_to_excel(self):
        # Load the existing Excel file or create a new one if it doesn't exist
        self.submit_label.pack()
        self.button1_next.pack_forget()
        self.button4_quit.pack()
        self.button3_submit.pack_forget()
        print(self.Image_Pair_No)
        print(self.Image_Score)
        self.add_values_to_excel(self.File_name, self.User_ID,self.Image_Pair_No, self.Image_Score, "Image_No","Image_Score")

    def add_values_to_excel(self, file_name, sheet_name, list1, list2, col1_name, col2_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
        except:
            workbook = openpyxl.Workbook()

        # Select the sheet or create a new one if it doesn't exist
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # Add column names to the first row
        worksheet.cell(row=1, column=1).value = col1_name
        worksheet.cell(row=1, column=2).value = col2_name

        # Append values from the first list to the first column
        for i in range(len(list1)):
            worksheet.cell(row=i + 2, column=1).value = list1[i]
        # Append values from the second list to the second column
        for i in range(len(list2)):
            worksheet.cell(row=i + 2, column=2).value = list2[i]

        # Save the changes to the Excel file
        workbook.save(file_name)

    def on_submit(self):
        dialog = tk.CTkInputDialog(text="Enter User ID:", title="User ID")
        user_ID = dialog.get_input()
        print((user_ID))


        if(user_ID!=""):
            self.User_ID_label.configure(self , text=f"User ID: {user_ID}")
            self.ctr_mid.grid(row=0, column=1, sticky="nsew")
            self.ctr_right.grid(row=0, column=2, sticky="nsew")
            self.checkbox_1.pack_forget()
            self.bottom_start.pack_forget()




jnd = JND_Survey()
jnd.mainloop()
